import os, glob
import sys
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
from pathlib import Path
from selenium import webdriver
# from seleniumwire import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
import time
from urllib.parse import urlparse
from selenium.webdriver.support.select import Select
import settings as s
import warnings
import requests
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
from html import unescape
import unicodedata
from slugify import slugify
from bs4 import BeautifulSoup
import ast

warnings.filterwarnings("ignore", category=UserWarning)
cookies = {
    'ab': 'stage-web-apps',
    '_ga': 'GA1.1.798189167.1703477051',
    '_gcl_au': '1.1.2118757811.1703477052',
    '_fbp': 'fb.1.1703732913826.708081806',
    '.AspNetCore.Antiforgery.VyLW6ORzMgk': 'CfDJ8KKvyj28jF5Ik7-myJ51XiTNoqvv0LQvrzQ1k3uaK30aHpfeJY64fIRQBqpvNif32vG0cWZgq5pBMZH-AgKEUj9zVbYGjLXFmHs8kAm6OEI0_GU__zXwxDElW_eng5EDQkTT_iBtgOsSrD42vgfC5Js',
    '_ga_YZFGTV3XRZ': 'GS1.1.1704604394.43.1.1704604408.46.0.0',
}
MAXROW=100
def createSheet(wb, no):
    newsheet = wb.create_sheet("PDF-" + str(no))
    newsheet.append(['VENDOR','MODEL ID', 'DIAGRAM ID', 'NAME','SECTION', 'DIAGRAM',	'PDF URL', 'LINK'])
    return newsheet

def parse():
    cud = s.CHROME_USER_DATA
    cp = s.CHROME_PROFILE
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir={}".format(cud))
    options.add_argument("profile-directory={}".format(cp))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    profile = {"download.default_directory": s.CHROME_DOWNLOAD_PATH + os.sep, 
                "download.extensions_to_open": "applications/pdf",
                "download.prompt_for_download": False,
                'profile.default_content_setting_values.automatic_downloads': 1,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True                   
                }
    options.add_experimental_option("prefs", profile)

    driver = webdriver.Chrome(service=Service(executable_path=os.path.join(os.getcwd(), "chromedriver", "chromedriver.exe")), options=options)
    driver.maximize_window()
    url = 'https://messicks.com/'
    driver.get(url)
    time.sleep(2)
    vendors = driver.find_elements(By.CSS_SELECTOR,"a[class='col-xs-4 col-sm-2 m-bottom-10 part-brand']")
    # breakpoint()
    vendorurls = []
    modelurls = []

    for vendor in vendors:
        vendorurl = vendor.get_attribute('href')
        vendorname = vendor.find_element(By.CSS_SELECTOR, "img").get_attribute('alt').replace('parts', "").strip()
        vendorurls.append((vendorurl, vendorname))

    # print(vendorurls)
    # sys.exit()
    input(vendorurls)
    for idx, vendorurl in enumerate(vendorurls):
        if vendorurl[1] != 'Befco':
            continue
        if os.path.exists(s.XLS_RESULT_PATH_V2 + os.sep +slugify(vendorurl[1]) + ".xlsx"):
            continue
        dlist = []
        no = 1
        wb = Workbook()
        ws = wb.active
        ws.title = 'Main-PDF'
        ws['A1'].value = "VENDOR"
        ws['B1'].value = "MODEL ID"
        ws['C1'].value = "MODEL NAME"
        ws['D1'].value = "URL" 
        ws['E1'].value = "ISDOWNLOAD"
        ws['F1'].value = "LINK"
        ws['G1'].value = "DESCRIPTION"

        # ws2 = wb.create_sheet("Sheet2")
        # ws2.append(['VENDOR','MODEL ID', 'DIAGRAM ID', 'NAME','SECTION', 'DIAGRAM',	'PDF URL', 'LINK'])

        headers = {
            'authority': 'messicks.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9,ja-JP;q=0.8,ja;q=0.7,id;q=0.6',
            'referer': f'{vendorurl[0]}',
            'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'x-requested-with': 'XMLHttpRequest',
        }


        driver.get(vendorurl[0])
        print(vendorurl[1], end="...", flush=True)
        time.sleep(2)
        brandId = driver.find_element(By.CSS_SELECTOR,"input#brandId").get_attribute('value')
        # print(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/0')
        response = requests.get(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/0', cookies=cookies, headers=headers)
        eqids1 = response.json()
        # breakpoint()
        vcount = 0
        for eqid1 in eqids1:
            eqidId1= eqid1['equipmentTypeId']
            # print(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/{eqidId1}')
            response = requests.get(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/{eqidId1}', cookies=cookies, headers=headers)
            eqids2 = response.json()
            if eqids2 == []:
                # print(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId1}')

                response = requests.get(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId1}', cookies=cookies, headers=headers)
                eqids2 = response.json()
                for eqid2 in eqids2:
                    theurl = f"https://messicks.com{eqid2['modelUrl']}"
                    # breakpoint()
                    modelname = unicodedata.normalize('NFKC',unescape(eqid2['modelName']))
                    dlist.append([vendorurl[1], str(int(eqid2['modelId'])), modelname, theurl, 'NO'])
                    no += 1
                    vcount += 1
            else:    
            # breakpoint()
                for eqid2 in eqids2:
                    eqidId2= eqid2['equipmentTypeId']
                    # print(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId2}')
                    response = requests.get(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId2}', cookies=cookies, headers=headers)
                    eqids3 = response.json()
                    for eqid3 in eqids3:
                        theurl = f"https://messicks.com{eqid3['modelUrl']}"
                        modelname = unicodedata.normalize('NFKC',unescape(eqid3['modelName']))
                        dlist.append([vendorurl[1], str(int(eqid3['modelId'])), modelname, theurl, 'NO'])
                        no += 1
                        vcount += 1
        print(vcount)
        for dt in dlist:
            ws.append(dt)
        tot = len(dlist)
        rownum = 0
        sheetnum = 1
        newsheet = createSheet(wb=wb, no=sheetnum)
        for idx, dt in enumerate(dlist):
            # if idx == 20:
            #     break
            print(idx+1, '-', tot, 'Extracting Diagrams for', dt[2], end="...", flush=True)
            response = requests.get(dt[3])
            soup = BeautifulSoup(response.text, "html.parser")
            script_tags = soup.find_all("script")
            jscript = ""
            for tag in script_tags:
                if tag.text.find("var _modelId") != -1:
                    jscript = tag.text
                    break
            secs = []
            diags = []
            modelid = ''
            for script in jscript.split("\n"):
                if "var _modelId" in script:
                    modelid = script.replace("var _modelId = '", "").replace("';","").strip()
                if "sections.push({" in script:
                    secs.append(script.replace("sectionId", "'sectionId'").replace("name :", "'name':").replace('sections.push(', '')[:-2])
                if "diagrams.push({" in script:
                    diags.append(script.replace("sectionId", "'sectionId'").replace("diagramId", "'diagramId'").replace("name :", "'name':").replace('diagrams.push(', '')[:-2])
            
            for sec in secs:
                try:
                    secdict =  ast.literal_eval(sec)
                except:
                    # secdict =  ast.literal_eval(unescape(sec))
                    secdict =  ast.literal_eval(sec.replace("\\","\\\\"))


                for diag in diags:
                    # breakpoint()
                    try:
                        diagdict = ast.literal_eval(diag)
                    except:
                        try:
                            diagdict = ast.literal_eval(diag.replace("\\","\\\\"))
                            # diagdict = ast.literal_eval(unescape(diag))
                        except:
                            continue
                    if secdict['sectionId'] == diagdict['sectionId']:
                        rownum += 1
                        dowloadurl = f"https://messicks.com/diagram/pdf?modelid={modelid}&diagramid={diagdict['diagramId']}"
                        newsheet.append([dt[0], modelid, diagdict['diagramId'], dt[0]+ " " + dt[2] + " Parts", unicodedata.normalize('NFKC',unescape(secdict['name'])), unicodedata.normalize('NFKC',unescape(diagdict['name']) ), dowloadurl])
                        if rownum == MAXROW:
                            # print('new sheet')
                            sheetnum += 1
                            rownum = 0
                            newsheet = createSheet(wb=wb, no=sheetnum)
            print("OK")

            # breakpoint()

        ws3 = wb.create_sheet("Setting")
        ws3.append(["file://<your_pdf_extract_location>"])
        ws3.append(["file://<your_pdf_join_location>"])
        print('Writing to file',slugify(vendorurl[1]) + ".xlsx",  end="...", flush=True)
        wb.save(s.XLS_RESULT_PATH_V2 + os.sep +slugify(vendorurl[1]) + ".xlsx")
        print("OK")
        # sys.exit()    
    driver.quit()


def main():
    parse()
    input("End Process..")    

if __name__ == '__main__':
    main()
