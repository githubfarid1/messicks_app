import os, glob
import sys
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
import time
from urllib.parse import urlparse
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from selenium.webdriver.support.select import Select
# import random
import settings as s
import argparse
import warnings
# import validators
# import pyexcel_ods3 as pods
from slugify import slugify
# import shutil
# from collections import OrderedDict
from html import unescape
import unicodedata
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
import xlwings as xw
from urllib.parse import urlparse
from urllib.request import urlretrieve
from openpyxl import Workbook, load_workbook

warnings.filterwarnings("ignore", category=UserWarning)

def genfilename(title, section, diagram):
    pathname = s.PDF_EXTRACT_PATH + os.sep + slugify("{}{}{}".format(title, section, diagram))
    if len(pathname) < 255:
        return pathname + ".pdf"
    else:
        while True:
            second = str(int(time.time()))
            newpathname = pathname[0:244] + second
            if os.path.exists(newpathname + ".pdf"):
                continue
            return newpathname + ".pdf"

def parse(modelid, datalist, xlbook):
    diagramlist = []
    # for sheet in xlsheets:
    modeldata = [item for item in datalist if item["modelid"] == modelid]
    # lastrow = xlsheet2.range('A' + str(xlsheet2.cells.last_cell.row)).end('up').row + 1
    success = 0 
    failed = 0
    # first = True
    # start = 0
    # end = 0
    # breakpoint()
    # for row in range(2, lastrow + 1):
    #     if first and xlsheet2[f'B{row}'].value == modelid:
    #         first = False
    #         start = row
    #     if not first and xlsheet2[f'B{row}'].value != modelid:
    #         end = row
    #         break
    
    for row in modeldata:
        # print(xlsheet2[f'B{row}'].value)
        dowloadurl = row['pdfurl']
        title, section, diagram  = row['name'], row['section'], row['diagram'].strip()
        pathname = genfilename(title, section, diagram)
        filename = os.path.basename(pathname)
        try:
            print("download for", title, section, diagram, end="... ", flush=True)
            if not os.path.exists(pathname):
                # breakpoint()
                urlretrieve(dowloadurl, pathname)
                print("OK")
            else:
                print("File Exists")
            link = '=HYPERLINK(CONCATENATE(Setting!$A$1,"{}"),"OPEN PDF")'.format(filename)
            diagramlist.append([section, diagram, link, dowloadurl, pathname])
            xlbook.sheets[f"PDF-{row['sheetnum']}"][f"H{row['rownum']}"].value = link
            success += 1

        except:
            xlbook.sheets[f"PDF-{row['sheetnum']}"][f"H{row['rownum']}"].value = 'NOT FOUND'
            failed += 1


    return diagramlist, success, failed

def main():
    parser = argparse.ArgumentParser(description="Catalog Product Downloader")
    parser.add_argument('-i', '--input', type=str,help="Source File")
    
    args = parser.parse_args()
    isExist = os.path.exists(args.input)
    if isExist == False :
        input('Please check the XLS file')
        sys.exit()

    source = args.input
    # driver = browser_init()
    # driver.maximize_window()
    print('Opening the Source Excel File...', end="", flush=True)
    # breakpoint()


    xlbook = xw.Book(source)
    xlsheet1 = xlbook.sheets["Main-PDF"]
    print("OK")
    print('Mount all data to memory...', end="", flush=True)
    time.sleep(1)
    pyxl = load_workbook(filename=source, read_only=True, data_only=True)
    sheetdata = []
    for sheetname in pyxl.sheetnames:
        if "PDF-" in sheetname:
            sheetdata.append(pyxl[sheetname])
    datalist = []
    for idx, sh in enumerate(sheetdata):
        first = True
        for rownum, row in enumerate(sh):
            if first:
                first = False
                continue
            mdict = {
                "rownum": rownum + 1,
                "sheetnum": idx + 1,
                "vendor": row[0].value,
                "modelid": row[1].value,
                "diagramid": row[2].value,
                "name": row[3].value,
                "section": row[4].value,
                "diagram": row[5].value,
                "pdfurl": row[6].value,
            }
            datalist.append(mdict)
    print("OK")

    # breakpoint()        
#-------------------------------------------    
    # xlbook = xw.Book(source)
    # xlsheet1 = xlbook.sheets["Main-PDF"]
    # xlsheets = [s for s in xlbook.sheets]
    # detailsheets = []
    # datalist = []
    # totalrow = 0
    # for sh in xlsheets:
    #     if "PDF-" in sh.name:
    #         detailsheets.append(sh)
    #         totalrow += sh.range('A' + str(sh.cells.last_cell.row)).end('up').row - 1
    # print("OK")
    # # breakpoint()
    # print('Mount all data to memory...')
    # time.sleep(1)
    # recno = 0
    # for idx, sh in enumerate(detailsheets):
    #     lastrow = sh.range('A' + str(sh.cells.last_cell.row)).end('up').row + 1
    #     for i in range(2, lastrow):
    #         recno += 1
    #         mdict = {
    #             "rownum": i,
    #             "sheetnum": idx + 1,
    #             "vendor": sh[f'A{i}'].value,
    #             "modelid": sh[f'B{i}'].value,
    #             "diagramid": sh[f'C{i}'].value,
    #             "name": sh[f'D{i}'].value,
    #             "section": sh[f'E{i}'].value,
    #             "diagram": sh[f'F{i}'].value,
    #             "pdfurl": sh[f'G{i}'].value,
    #         }
    #         print('mount record', recno, '-', totalrow)
    #         datalist.append(mdict)
    # # breakpoint()
               
    # # xlsheet2 = xlbook.sheets["Sheet2"]
    # print('OK')
    # --------------------------------
    maxrow = xlsheet1.range('C' + str(xlsheet1.cells.last_cell.row)).end('up').row
    # breakpoint()
    for i in range(2, maxrow + 1):
        merger = PdfMerger()
        vendor = xlsheet1[f'A{i}'].value
        title = vendor + " " + xlsheet1[f'C{i}'].value + " Parts"
        if xlsheet1[f'E{i}'].value == 'NO':
            # print('search',xlsheet1[f'B{i}'].value)
            diagramlist, success, failed = parse(modelid=xlsheet1[f'B{i}'].value, datalist=datalist, xlbook=xlbook)
            # break
            # diagramlist.append([section, diagram, link, dowloadurl, pathname])
            if len(diagramlist) > 0:
                for diagram in diagramlist:
                    filename = str(diagram[2]).split(",")[1].replace('"',"").replace(")","")
                    try:
                        merger.append(s.PDF_EXTRACT_PATH + os.sep + filename)
                        # print("merge", filename)
                    except:
                        urlretrieve(diagram[3], diagram[4])
                        merger.append(s.PDF_EXTRACT_PATH + os.sep + filename)
                        # breakpoint()

                merger.write(s.PDF_JOIN_PATH + os.sep + slugify(title) + ".pdf")
                link = '=HYPERLINK(CONCATENATE(Setting!$A$2,"{}"),"OPEN PDF")'.format(slugify(title) + ".pdf")
                xlsheet1[f'E{i}'].value = 'YES'
                xlsheet1[f'F{i}'].value = link
                xlsheet1[f'G{i}'].value = f"PDF Download Success={success}, Failed={failed}"
            else:
                xlsheet1[f'E{i}'].value = 'YES'
                xlsheet1[f'F{i}'].value = "FAILED"
                xlsheet1[f'G{i}'].value = "No PDF can not be Downloaded"
    # driver.quit()


    input("End Process..")    


if __name__ == '__main__':
    main()
