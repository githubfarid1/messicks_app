import os, glob
import sys
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
from pathlib import Path
from selenium import webdriver
# from seleniumwire import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager as CM
import time
from urllib.parse import urlparse
from selenium.webdriver.support.select import Select
import settings as s
import warnings
import requests
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
from html import unescape
import unicodedata
from slugify import slugify
from bs4 import BeautifulSoup


warnings.filterwarnings("ignore", category=UserWarning)
# cookies = {
#     'ab': 'stage-web-apps',
#     '_ga': 'GA1.1.798189167.1703477051',
#     '_gcl_au': '1.1.2118757811.1703477052',
#     '_fbp': 'fb.1.1703732913826.708081806',
#     # '.AspNetCore.Antiforgery.VyLW6ORzMgk': 'CfDJ8KKvyj28jF5Ik7-myJ51XiThun0j6OjyoQeyuImmlypZedyWoBRLRjadyzu2ReoTaLVAFNiJeD-mg6qnDBvXnTzYkHQrkTGXz5wWmIZ58ihXBNDY6jE2dlLGmTCzQDne443QVhFeSwsPGs_HMlpOmto',
#     '_ga_YZFGTV3XRZ': 'GS1.1.1704154533.23.1.1704157025.6.0.0',
# }
# cookies = {
#     'ab': 'stage-web-apps',
#     '_ga': 'GA1.1.798189167.1703477051',
#     '_gcl_au': '1.1.2118757811.1703477052',
#     '_fbp': 'fb.1.1703732913826.708081806',
#     '.AspNetCore.Antiforgery.VyLW6ORzMgk': 'CfDJ8KKvyj28jF5Ik7-myJ51XiRff1tD0FFAu_HUpKxyMUIALdH8ZD3PODjXOPJOjCLVYU77bZQo0HW_s9WsjmgMG3UElP-gvUeG6K5sBKP2GC-_sOTryRgMMbI9CFdXnLqCeqlWhY7P84Kfz5sqesxP3R8',
#     '_ga_YZFGTV3XRZ': 'GS1.1.1704344393.35.1.1704344415.38.0.0',
# }
cookies = {
    'ab': 'stage-web-apps',
    '_ga': 'GA1.1.798189167.1703477051',
    '_gcl_au': '1.1.2118757811.1703477052',
    '_fbp': 'fb.1.1703732913826.708081806',
    '.AspNetCore.Antiforgery.VyLW6ORzMgk': 'CfDJ8KKvyj28jF5Ik7-myJ51XiTNoqvv0LQvrzQ1k3uaK30aHpfeJY64fIRQBqpvNif32vG0cWZgq5pBMZH-AgKEUj9zVbYGjLXFmHs8kAm6OEI0_GU__zXwxDElW_eng5EDQkTT_iBtgOsSrD42vgfC5Js',
    '_ga_YZFGTV3XRZ': 'GS1.1.1704604394.43.1.1704604408.46.0.0',
}

def parse():
    cud = s.CHROME_USER_DATA
    cp = s.CHROME_PROFILE
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir={}".format(cud))
    options.add_argument("profile-directory={}".format(cp))
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--window-size=800,600")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    profile = {"download.default_directory": s.CHROME_DOWNLOAD_PATH + os.sep, 
                "download.extensions_to_open": "applications/pdf",
                "download.prompt_for_download": False,
                'profile.default_content_setting_values.automatic_downloads': 1,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True                   
                }
    options.add_experimental_option("prefs", profile)

    driver = webdriver.Chrome(service=Service(executable_path=os.path.join(os.getcwd(), "chromedriver", "chromedriver.exe")), options=options)
    driver.maximize_window()
    url = 'https://messicks.com/'
    driver.get(url)
    time.sleep(2)
    vendors = driver.find_elements(By.CSS_SELECTOR,"a[class='col-xs-4 col-sm-2 m-bottom-10 part-brand']")
    # breakpoint()
    vendorurls = []
    modelurls = []

    for vendor in vendors:
        vendorurl = vendor.get_attribute('href')
        vendorname = vendor.find_element(By.CSS_SELECTOR, "img").get_attribute('alt').replace('parts', "").strip()
        vendorurls.append((vendorurl, vendorname))

    # print(vendorurls)
    # sys.exit()
    for idx, vendorurl in enumerate(vendorurls):
        # if idx != 2:
        #     continue
        dlist = []

        no = 1
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws['A1'].value = "VENDOR"
        ws['B1'].value = "MODEL NAME"
        ws['C1'].value = "URL" 
        ws['D1'].value = "ISDOWNLOAD"
        ws['E1'].value = "LINK"
        ws['F1'].value = "DESCRIPTION"

        headers = {
            'authority': 'messicks.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9,ja-JP;q=0.8,ja;q=0.7,id;q=0.6',
            # 'cookie': 'ab=stage-web-apps; _ga=GA1.1.798189167.1703477051; _gcl_au=1.1.2118757811.1703477052; _fbp=fb.1.1703732913826.708081806; .AspNetCore.Antiforgery.VyLW6ORzMgk=CfDJ8KKvyj28jF5Ik7-myJ51XiThun0j6OjyoQeyuImmlypZedyWoBRLRjadyzu2ReoTaLVAFNiJeD-mg6qnDBvXnTzYkHQrkTGXz5wWmIZ58ihXBNDY6jE2dlLGmTCzQDne443QVhFeSwsPGs_HMlpOmto; _ga_YZFGTV3XRZ=GS1.1.1704154533.23.1.1704157025.6.0.0',
            'referer': f'{vendorurl[0]}',
            'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'x-requested-with': 'XMLHttpRequest',
        }


        driver.get(vendorurl[0])
        print(vendorurl[1], end="...", flush=True)
        time.sleep(2)
        brandId = driver.find_element(By.CSS_SELECTOR,"input#brandId").get_attribute('value')
        # print(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/0')
        response = requests.get(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/0', cookies=cookies, headers=headers)
        eqids1 = response.json()
        # breakpoint()
        vcount = 0
        for eqid1 in eqids1:
            eqidId1= eqid1['equipmentTypeId']
            # print(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/{eqidId1}')
            response = requests.get(f'https://messicks.com/api/vendor/equipmenttypes/{brandId}/{eqidId1}', cookies=cookies, headers=headers)
            eqids2 = response.json()
            if eqids2 == []:
                # print(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId1}')

                response = requests.get(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId1}', cookies=cookies, headers=headers)
                eqids2 = response.json()
                for eqid2 in eqids2:
                    theurl = f"https://messicks.com{eqid2['modelUrl']}"
                    modelname = unicodedata.normalize('NFKC',unescape(eqid2['modelName']))
                    dlist.append([vendorurl[1], modelname, theurl, 'NO'])
                    no += 1
                    vcount += 1
            else:    
            # breakpoint()
                for eqid2 in eqids2:
                    eqidId2= eqid2['equipmentTypeId']
                    # print(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId2}')
                    response = requests.get(f'https://messicks.com/api/vendor/models/{brandId}/{eqidId2}', cookies=cookies, headers=headers)
                    eqids3 = response.json()
                    for eqid3 in eqids3:
                        theurl = f"https://messicks.com{eqid3['modelUrl']}"
                        modelname = unicodedata.normalize('NFKC',unescape(eqid3['modelName']))
                        dlist.append([vendorurl[1], modelname, theurl, 'NO'])
                        no += 1
                        vcount += 1
        print(vcount)
        for dt in dlist:
            ws.append(dt)


        ws2 = wb.create_sheet("Sheet2")
        ws2.append(['VENDOR','NAME','SECTION',	'DIAGRAM',	'LINK'])
        ws3 = wb.create_sheet("Sheet3")
        ws3.append(["file://<your_pdf_extract_location>"])
        ws3.append(["file://<your_pdf_join_location>"])
        wb.save(s.XLS_RESULT_PATH + os.sep +slugify(vendorurl[1]) + ".xlsx")
    
    driver.quit()


def main():
    parse()
    input("End Process..")    

if __name__ == '__main__':
    main()
