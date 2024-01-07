from collections import OrderedDict
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import pyexcel_ods3 as pods
odssource = "odfs_result\\resulturls.ods"
xlssource = "odfs_result\\resulturls.xlsx"
odswb = pods.get_data(afile=odssource)
odssheet1 = odswb['Sheet1'].copy()
odssheet2 = odswb['Sheet2'].copy()
odssheet3 = odswb['Sheet3'].copy()

datadict = OrderedDict()
xlswb = load_workbook(filename = xlssource)
xlssheet1 = xlswb['Sheet1']
xlslist = []
for i in range(2, xlssheet1.max_row + 1):
    xlslist.append((xlssheet1[f"B{i}"].value, xlssheet1[f"C{i}"].value))

for idx, data in enumerate(odssheet1):
    if idx == 0:
        continue
    modelname = ""
    for xldata in xlslist:
        if data[1] == xldata[1]:
            modelname = xldata[0]
            break
    # breakpoint()
    try:
        odssheet1[idx] = [data[0], modelname, data[1], data[2], data[3]]
    except:
        odssheet1[idx] = [data[0], modelname, data[1], data[2]]
datadict.update({"Sheet1": odssheet1})
datadict.update({"Sheet2": odssheet2})
datadict.update({"Sheet3": odssheet3})
pods.save_data("odfs_result\\resulturls_new.ods", datadict)